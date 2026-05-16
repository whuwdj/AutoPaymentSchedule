# -*- coding: utf-8 -*-
"""
按 Sheet4「负责人→成本名称」顺序，从 Sheet1 汇总「付款计划」列：
- 付款优先级 = 0：按负责人汇总（nCountZero*_0）及板块合计 sum_zero；
- 付款优先级 = 1 / 2：按负责人汇总（nCountZero*_1 / *_2）及合计 sum_priority1 / sum_priority2；
- 全量：不按优先级筛选，按负责人汇总（nCountAll*）及合计 sum_all。
查看排款前置：Sheet1「付款计划」列须存在至少一格可解析数值，否则提示先智能排款。
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from excel_service import (
    _cell_text_normalized,
    _coerce_numeric_for_sum,
    _is_xls,
    _norm_header_label,
)
from smart_schedule import (
    _find_col_header_contains,
    _find_payment_plan_col,
    _find_sheet1_header_row,
    _priority_level_0_2,
)


@dataclass(frozen=True)
class OwnerCostPanelStats:
    """与 Sheet4 负责人首次出现顺序一致；展示用标签为成本名称（空则回退负责人）。"""

    cost_labels: Tuple[str, ...]
    per_owner_zero: Tuple[float, ...]
    per_owner_priority1: Tuple[float, ...]
    per_owner_priority2: Tuple[float, ...]
    per_owner_all: Tuple[float, ...]
    sum_zero: float
    sum_priority1: float
    sum_priority2: float
    sum_all: float


def _find_sheet4_header_cols(ws: Worksheet) -> Optional[Tuple[int, int, int]]:
    """返回 (表头行, 负责人列, 成本名称列)，均为 1-based。"""
    max_r = min(ws.max_row or 0, 50)
    max_c = min(ws.max_column or 0, 80)
    for r in range(1, max_r + 1):
        c_owner: Optional[int] = None
        c_cost: Optional[int] = None
        for c in range(1, max_c + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            tn = _norm_header_label(_cell_text_normalized(cell.value))
            if tn == "负责人":
                c_owner = c
            elif tn == "成本名称":
                c_cost = c
        if c_owner is not None and c_cost is not None:
            return r, c_owner, c_cost
    return None


def _sheet4_ordered_owner_cost_pairs(
    ws: Worksheet, header_row: int, c_owner: int, c_cost: int
) -> List[Tuple[str, str]]:
    """按表顺序；同一负责人仅保留首次出现。"""
    seen: set[str] = set()
    out: List[Tuple[str, str]] = []
    for r in range(header_row + 1, (ws.max_row or 0) + 1):
        co = ws.cell(row=r, column=c_owner)
        if isinstance(co, MergedCell):
            continue
        owner = _cell_text_normalized(co.value)
        if not owner:
            continue
        if owner in seen:
            continue
        seen.add(owner)
        cc = ws.cell(row=r, column=c_cost)
        cost_nm = ""
        if not isinstance(cc, MergedCell):
            cost_nm = _cell_text_normalized(cc.value)
        out.append((owner, cost_nm))
    return out


def _find_sheet1_owner_col(ws: Worksheet, header_row: int) -> Optional[int]:
    """优先「负责人」，否则「主体」（表头可带中英文冒号）。"""
    max_c = ws.max_column or 0
    for want in ("负责人", "主体"):
        for c in range(1, max_c + 1):
            cell = ws.cell(row=header_row, column=c)
            if isinstance(cell, MergedCell):
                continue
            t = _norm_header_label(_cell_text_normalized(cell.value))
            if t == want:
                return c
    return None


def load_sheet4_cost_display_labels(path: str) -> Optional[Tuple[str, ...]]:
    """
    仅从 Sheet4 读取负责人顺序与展示标签（成本名称，空则回退负责人）。
    用于无汇总数据时仍与 Sheet4 列顺序、标签一致地占位 0.00。
    """
    if _is_xls(path):
        return None
    wb = load_workbook(path, read_only=False, data_only=True)
    try:
        if len(wb.worksheets) < 4:
            return None
        ws4 = wb.worksheets[3]
        h4 = _find_sheet4_header_cols(ws4)
        if h4 is None:
            return None
        hr4, c4o, c4c = h4
        pairs = _sheet4_ordered_owner_cost_pairs(ws4, hr4, c4o, c4c)
        if not pairs:
            return None
        return tuple(
            (cost_nm if cost_nm else owner) for owner, cost_nm in pairs
        )
    finally:
        wb.close()


def sheet1_payment_plan_has_numeric_entries(path: str) -> bool:
    """
    Sheet1 数据区「付款计划」列是否存在至少一格可解析为数值（含 0）。
    若无，视为「付款计划列全部为空」；用于「查看排款」前置。
    """
    if _is_xls(path):
        return False
    wb = load_workbook(path, read_only=False, data_only=True)
    try:
        ws1 = wb.worksheets[0]
        hr1 = _find_sheet1_header_row(ws1)
        if hr1 is None:
            return False
        c_plan = _find_payment_plan_col(ws1, hr1)
        if c_plan is None:
            return False
        max_r1 = ws1.max_row or 0
        for r in range(hr1 + 1, max_r1 + 1):
            cell_p = ws1.cell(row=r, column=c_plan)
            if isinstance(cell_p, MergedCell):
                continue
            if _coerce_numeric_for_sum(cell_p.value) is not None:
                return True
        return False
    finally:
        wb.close()


def load_owner_cost_panel_stats(path: str) -> Optional[OwnerCostPanelStats]:
    """
    读取总表：Sheet4 定顺序与成本名称；Sheet1 按负责人汇总付款计划列。
    非 .xlsx、缺表/缺列时返回 None。
    """
    if _is_xls(path):
        return None
    wb = load_workbook(path, read_only=False, data_only=True)
    try:
        if len(wb.worksheets) < 4:
            return None
        ws4 = wb.worksheets[3]
        h4 = _find_sheet4_header_cols(ws4)
        if h4 is None:
            return None
        hr4, c4o, c4c = h4
        pairs = _sheet4_ordered_owner_cost_pairs(ws4, hr4, c4o, c4c)
        if not pairs:
            return None

        ws1 = wb.worksheets[0]
        hr1 = _find_sheet1_header_row(ws1)
        if hr1 is None:
            return None
        c_owner = _find_sheet1_owner_col(ws1, hr1)
        if c_owner is None:
            return None
        c_pri = _find_col_header_contains(ws1, hr1, "付款优先级")
        if c_pri is None:
            return None
        c_plan = _find_payment_plan_col(ws1, hr1)
        if c_plan is None:
            return None

        sum_zero: Dict[str, float] = {}
        sum_p1: Dict[str, float] = {}
        sum_p2: Dict[str, float] = {}
        sum_all: Dict[str, float] = {}
        max_r1 = ws1.max_row or 0
        for r in range(hr1 + 1, max_r1 + 1):
            cell_o = ws1.cell(row=r, column=c_owner)
            if isinstance(cell_o, MergedCell):
                continue
            owner = _cell_text_normalized(cell_o.value)
            if not owner:
                continue
            cell_p = ws1.cell(row=r, column=c_plan)
            if isinstance(cell_p, MergedCell):
                continue
            n = _coerce_numeric_for_sum(cell_p.value)
            amt = float(n) if n is not None else 0.0
            sum_all[owner] = sum_all.get(owner, 0.0) + amt
            pl = _priority_level_0_2(ws1.cell(row=r, column=c_pri).value)
            if pl == 0:
                sum_zero[owner] = sum_zero.get(owner, 0.0) + amt
            elif pl == 1:
                sum_p1[owner] = sum_p1.get(owner, 0.0) + amt
            elif pl == 2:
                sum_p2[owner] = sum_p2.get(owner, 0.0) + amt

        cost_labels: List[str] = []
        zero_list: List[float] = []
        p1_list: List[float] = []
        p2_list: List[float] = []
        all_list: List[float] = []
        for owner, cost_nm in pairs:
            label = cost_nm if cost_nm else owner
            cost_labels.append(label)
            zero_list.append(sum_zero.get(owner, 0.0))
            p1_list.append(sum_p1.get(owner, 0.0))
            p2_list.append(sum_p2.get(owner, 0.0))
            all_list.append(sum_all.get(owner, 0.0))

        return OwnerCostPanelStats(
            cost_labels=tuple(cost_labels),
            per_owner_zero=tuple(zero_list),
            per_owner_priority1=tuple(p1_list),
            per_owner_priority2=tuple(p2_list),
            per_owner_all=tuple(all_list),
            sum_zero=sum(zero_list),
            sum_priority1=sum(p1_list),
            sum_priority2=sum(p2_list),
            sum_all=sum(all_list),
        )
    finally:
        wb.close()
