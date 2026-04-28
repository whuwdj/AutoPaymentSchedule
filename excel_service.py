# -*- coding: utf-8 -*-
"""排款计划总表 Sheet2：读取银行列表、写入可支付金额与付款方式。.xlsx 完整支持；.xls 仅支持读取银行列表。"""

from __future__ import annotations

import os
import re
from dataclasses import dataclass
from datetime import date, datetime
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# xlrd 1.2.0 用于 .xls 只读
import xlrd


EXCEL_FILTER = "Excel (*.xlsx *.xls);;Excel 2007+ (*.xlsx);;Excel 97-2003 (*.xls)"

# 首表（Sheet1）横向摘要行：五个表头词须同现一行，再各自向右取到首个空单元格
SHEET1_SUMMARY_LABELS: Tuple[str, ...] = (
    "银行名称",
    "排款总额",
    "放款日期",
    "可支付的主体",
    "付款方式",
)


@dataclass(frozen=True)
class Sheet1SummaryStripResult:
    """首表命中行及每个表头右侧连续非空单元格（遇空或遇下一表头词则停）。"""

    row_1based: Optional[int]
    columns: Dict[str, Tuple[str, ...]]


@dataclass(frozen=True)
class BankOption:
    """下拉项：[序号] 银行名称"""

    serial: int
    name: str
    row_index_1based: int  # Sheet 内数据行号（含表头则为 Excel 行号）

    def display(self) -> str:
        return f"[{self.serial}] {self.name}"


def _is_xls(path: str) -> bool:
    p = path.lower()
    if p.endswith(".xlsx"):
        return False
    return p.endswith(".xls")


def find_latest_total_excel(directory: str) -> Optional[str]:
    if not os.path.isdir(directory):
        return None
    candidates: List[str] = []
    for name in os.listdir(directory):
        lower = name.lower()
        if lower.endswith(".xlsx"):
            candidates.append(os.path.join(directory, name))
        elif lower.endswith(".xls"):
            candidates.append(os.path.join(directory, name))
    if not candidates:
        return None
    return max(candidates, key=lambda p: os.path.getmtime(p))


def _scan_header(ws, max_scan_rows: int = 10) -> Tuple[int, dict]:
    """
    返回 (表头所在行号 1-based, {关键字: 列号 1-based})
    关键字: serial, bank, amount, method
    """
    keys = {
        "serial": ("序号",),
        "bank": ("银行名称", "银行", "开户银行"),
        "amount": ("可支付金额", "支付金额", "金额"),
        "method": ("付款方式", "支付方式"),
    }
    found: dict = {}
    best_row = 1
    max_c = ws.max_column or 1
    for r in range(1, max_scan_rows + 1):
        row_map: dict = {}
        for c in range(1, max_c + 1):
            cell = ws.cell(row=r, column=c).value
            if cell is None:
                continue
            text = str(cell).strip()
            for key, variants in keys.items():
                if key in row_map:
                    continue
                for v in variants:
                    if v in text or text == v:
                        row_map[key] = c
                        break
        # 至少要有 序号 + 银行
        if "serial" in row_map and "bank" in row_map:
            found = row_map
            best_row = r
            break
    return best_row, found


def read_banks_from_xlsx(path: str) -> Tuple[List[BankOption], str]:
    wb = load_workbook(path, read_only=False, data_only=True)
    if len(wb.worksheets) < 2:
        wb.close()
        raise ValueError("排款计划总表第二页（Sheet2）不存在。")
    ws = wb.worksheets[1]
    header_row, colmap = _scan_header(ws)
    if "serial" not in colmap or "bank" not in colmap:
        wb.close()
        raise ValueError("未在 Sheet2 表头中找到「序号」与「银行名称」列，请检查总表格式。")
    c_serial = colmap["serial"]
    c_bank = colmap["bank"]
    out: List[BankOption] = []
    for r in range(header_row + 1, ws.max_row + 1):
        sval = ws.cell(row=r, column=c_serial).value
        bval = ws.cell(row=r, column=c_bank).value
        if bval is None or str(bval).strip() == "":
            continue
        try:
            serial = int(float(sval)) if sval is not None else r - header_row
        except (TypeError, ValueError):
            serial = r - header_row
        name = str(bval).strip()
        out.append(BankOption(serial=serial, name=name, row_index_1based=r))
    wb.close()
    if not out:
        raise ValueError("Sheet2 中未读取到任何银行行数据。")
    return out, os.path.basename(path)


def read_banks_from_xls(path: str) -> Tuple[List[BankOption], str]:
    book = xlrd.open_workbook(path)
    if book.nsheets < 2:
        raise ValueError("排款计划总表第二页（Sheet2）不存在。")
    sh = book.sheet_by_index(1)
    # 扫描表头
    header_row = 0
    colmap: dict = {}
    keys = {
        "serial": ("序号",),
        "bank": ("银行名称", "银行", "开户银行"),
    }
    for r in range(min(10, sh.nrows)):
        row_map: dict = {}
        for c in range(sh.ncols):
            cell = sh.cell_value(r, c)
            if cell is None or cell == "":
                continue
            text = str(cell).strip()
            for key, variants in keys.items():
                if key in row_map:
                    continue
                for v in variants:
                    if v in text or text == v:
                        row_map[key] = c
                        break
        if "serial" in row_map and "bank" in row_map:
            colmap = row_map
            header_row = r
            break
    if "serial" not in colmap or "bank" not in colmap:
        raise ValueError("未在 Sheet2 表头中找到「序号」与「银行名称」列，请检查总表格式。")
    c_serial = colmap["serial"]
    c_bank = colmap["bank"]
    out: List[BankOption] = []
    for r in range(header_row + 1, sh.nrows):
        sval = sh.cell_value(r, c_serial)
        bval = sh.cell_value(r, c_bank)
        if bval is None or str(bval).strip() == "":
            continue
        try:
            serial = int(float(sval)) if sval != "" else r - header_row
        except (TypeError, ValueError):
            serial = r - header_row
        name = str(bval).strip()
        out.append(BankOption(serial=serial, name=name, row_index_1based=r + 1))  # xlrd 0-based -> 1-based Excel row
    if not out:
        raise ValueError("Sheet2 中未读取到任何银行行数据。")
    return out, os.path.basename(path)


def read_banks(path: str) -> Tuple[List[BankOption], str]:
    if _is_xls(path):
        return read_banks_from_xls(path)
    return read_banks_from_xlsx(path)


def _cell_text_normalized(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S").strip()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _collect_strip_openpyxl(
    ws: Worksheet,
    row_1based: int,
    label_col: int,
    label_values: frozenset[str],
) -> List[str]:
    out: List[str] = []
    max_c = ws.max_column or 0
    c = label_col + 1
    while c <= max_c:
        t = _cell_text_normalized(ws.cell(row=row_1based, column=c).value)
        if not t:
            break
        if t in label_values:
            break
        out.append(t)
        c += 1
    return out


def read_sheet1_summary_strip_xlsx(path: str) -> Sheet1SummaryStripResult:
    wb = load_workbook(path, read_only=False, data_only=True)
    try:
        ws = wb.worksheets[0]
        labels = SHEET1_SUMMARY_LABELS
        label_froze = frozenset(labels)
        max_scan_r = min(ws.max_row or 0, 500)
        max_c = ws.max_column or 0
        best_row: Optional[int] = None
        best_cols: Dict[str, int] = {}
        for r in range(1, max_scan_r + 1):
            cols_by_label: Dict[str, int] = {}
            for c in range(1, max_c + 1):
                t = _cell_text_normalized(ws.cell(row=r, column=c).value)
                if t in labels and t not in cols_by_label:
                    cols_by_label[t] = c
            if len(cols_by_label) == len(labels):
                best_row = r
                best_cols = cols_by_label
                break
        if best_row is None:
            return Sheet1SummaryStripResult(row_1based=None, columns={lb: () for lb in labels})
        cols_data: Dict[str, Tuple[str, ...]] = {}
        for lb in labels:
            cols_data[lb] = tuple(
                _collect_strip_openpyxl(ws, best_row, best_cols[lb], label_froze)
            )
        return Sheet1SummaryStripResult(row_1based=best_row, columns=cols_data)
    finally:
        wb.close()


def read_sheet1_summary_strip_xls(path: str) -> Sheet1SummaryStripResult:
    book = xlrd.open_workbook(path)
    sh = book.sheet_by_index(0)
    labels = SHEET1_SUMMARY_LABELS
    label_froze = frozenset(labels)
    max_scan_r = min(sh.nrows, 500)
    max_c = sh.ncols

    def cell_str(r: int, c0: int) -> str:
        try:
            v = sh.cell_value(r, c0)
        except IndexError:
            return ""
        if v == "" or v is None:
            return ""
        tp = sh.cell_type(r, c0)
        if tp == xlrd.XL_CELL_DATE:
            try:
                y, mo, d, h, mi, s = xlrd.xldate_as_tuple(v, book.datemode)
                if h == 0 and mi == 0 and s == 0:
                    return f"{y:04d}-{mo:02d}-{d:02d}"
                return f"{y:04d}-{mo:02d}-{d:02d} {h:02d}:{mi:02d}:{s:02d}"
            except Exception:
                return str(v).strip()
        return str(v).strip()

    best_r0: Optional[int] = None
    best_cols_1based: Dict[str, int] = {}
    for r in range(max_scan_r):
        cols_by_label: Dict[str, int] = {}
        for c0 in range(max_c):
            t = cell_str(r, c0)
            if t in labels and t not in cols_by_label:
                cols_by_label[t] = c0 + 1
        if len(cols_by_label) == len(labels):
            best_r0 = r
            best_cols_1based = cols_by_label
            break

    if best_r0 is None:
        return Sheet1SummaryStripResult(row_1based=None, columns={lb: () for lb in labels})

    cols_data: Dict[str, Tuple[str, ...]] = {}
    for lb in labels:
        c1 = best_cols_1based[lb]
        out: List[str] = []
        c = c1 + 1
        while c <= max_c:
            t = cell_str(best_r0, c - 1)
            if not t:
                break
            if t in label_froze:
                break
            out.append(t)
            c += 1
        cols_data[lb] = tuple(out)
    return Sheet1SummaryStripResult(row_1based=best_r0 + 1, columns=cols_data)


def read_sheet1_summary_strip(path: str) -> Sheet1SummaryStripResult:
    """读取首表 Sheet1 上同时含五个关键词的一行，并横向截取各词右侧直到空的片段。"""
    if _is_xls(path):
        return read_sheet1_summary_strip_xls(path)
    return read_sheet1_summary_strip_xlsx(path)


def _ensure_amount_method_cols(ws: Worksheet, header_row: int, colmap: dict) -> dict:
    """若表头无金额/方式列，在表尾追加两列并写表头。"""
    out = dict(colmap)
    next_col = ws.max_column + 1
    if "amount" not in out:
        ws.cell(row=header_row, column=next_col, value="可支付金额")
        out["amount"] = next_col
        next_col += 1
    if "method" not in out:
        ws.cell(row=header_row, column=next_col, value="付款方式")
        out["method"] = next_col
    return out


def write_bank_config_xlsx(
    path: str,
    bank: BankOption,
    amount: float,
    payment_method: str,
) -> None:
    wb = load_workbook(path, read_only=False, data_only=False)
    if len(wb.worksheets) < 2:
        wb.close()
        raise ValueError("排款计划总表第二页（Sheet2）不存在。")
    ws = wb.worksheets[1]
    header_row, colmap = _scan_header(ws)
    if "serial" not in colmap or "bank" not in colmap:
        wb.close()
        raise ValueError("未在 Sheet2 表头中找到「序号」与「银行名称」列。")
    colmap = _ensure_amount_method_cols(ws, header_row, colmap)
    c_serial = colmap["serial"]
    c_bank = colmap["bank"]
    c_amount = colmap["amount"]
    c_method = colmap["method"]

    target_row: Optional[int] = None
    for r in range(header_row + 1, ws.max_row + 1):
        sval = ws.cell(row=r, column=c_serial).value
        bval = ws.cell(row=r, column=c_bank).value
        if bval is None or str(bval).strip() == "":
            continue
        try:
            ser = int(float(sval)) if sval is not None else None
        except (TypeError, ValueError):
            ser = None
        name = str(bval).strip()
        if ser == bank.serial and name == bank.name:
            target_row = r
            break
        if name == bank.name and (ser is None or ser == bank.serial):
            target_row = r
            break
    if target_row is None:
        wb.close()
        raise ValueError(f"未在 Sheet2 中找到银行：{bank.display()}")

    ws.cell(row=target_row, column=c_amount, value=amount)
    ws.cell(row=target_row, column=c_method, value=payment_method)
    wb.save(path)
    wb.close()


def parse_display_choice(text: str) -> Optional[Tuple[int, str]]:
    """从 '[1] 中国银行' 解析 (1, '中国银行')"""
    m = re.match(r"^\s*\[(\d+)\]\s*(.+)\s*$", text.strip())
    if not m:
        return None
    return int(m.group(1)), m.group(2).strip()
