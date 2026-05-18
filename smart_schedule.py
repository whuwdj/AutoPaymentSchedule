# -*- coding: utf-8 -*-
"""
智能排款：读取总表 Sheet1（程序目录 AutoPaymentScheduleFile/TotalSheet 下 Excel），
初始化表后按付款优先级 0→1→2 依次处理；同档内按 Sheet2「类别优先级」+ Sheet2「序号」+ Sheet1「序号」排序，
类别未匹配行排在同档末尾并按 Sheet1「序号」稳定排序；尝试各「笔」列时仅调整遍历顺序（不改列位置）：
按 Sheet3 银行应付款/付款优先级数值升序，未配置银行排最后；账期与应收汇总、nPeriodPercent 着色、写回「笔」列与付款计划并扣减第 5 行排款余额。

核心计算（与业务规格一致，见 run_smart_schedule_on_total_sheet / _compute_ntotal_for_priority）：
- 账期数 nPeriodCount = 协议账期÷30（_protocol_period_months）。
- nPayablesSum：「当月已开票余额」列之后～「截止」列之前，表头 nMonth（「N个月」取 N）满足 nMonth≥nPeriodCount
  且单元格数值≥0 的格求和（_sum_over_invoice_cutoff_window，负数不计）；截止列及其右侧不参与。
- 优先级 0：nTotalSum=nPayablesSum；上述区间内数值非空且>0 的格标浅绿（_highlight_priority0_month_window）。
- 优先级 1/2：nTotalSum=nPayablesSum×(界面整数百分比÷100)；锚点为上述区间内自右向左最后一个数值非空列，
  nPeriodPercent=(锚点值/nPayablesSum)×100，与 nPercentOne/nPercentTwo 比较后循环降月或分格渐变着色
  （_compute_ntotal_priority12_with_highlights）。
- 排款完成后卡片 2：严格按账期排款额=各优先级 0 行 nTotalSum 之和；第一/二优先级排款额=「付款计划」列中
  该行「付款优先级」为 1/2 的金额之和；本月合计排款总额=三者之和（见 SmartScheduleResult）。

上传总表后界面「应付额」四行可由 summarize_card1_payable_from_total_sheet 按账期窗口与行优先级预填。
"""

from __future__ import annotations

import math
import re
from collections import defaultdict
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from typing import Dict, List, Optional, Tuple, Union

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.styles import GradientFill, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.styles.fills import Stop
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from openpyxl.worksheet.worksheet import Worksheet

from excel_service import (
    _cell_text_normalized,
    _coerce_numeric_for_sum,
    _is_xls,
    _norm_header_label,
    sum_sheet1_row4_after_payout_through_total_scheduled,
)


@dataclass
class SmartScheduleResult:
    path: str
    rows_priority0: int
    rows_priority1: int
    rows_priority2: int
    rows_scanned: int
    rows_written: int
    rows_skipped: int
    reserved_total: Optional[float] = None
    priority0_total: float = 0.0
    priority1_total: float = 0.0
    priority2_total: float = 0.0


@dataclass(frozen=True)
class Card1PayableTotals:
    """卡片 1 应付额：按账期窗口逐行汇总后按付款优先级 0/1/2 累计，及总应付 nCountTotal。"""

    n_count0: float
    n_count1: float
    n_count2: float
    n_count_total: float


def _norm_h(text: str) -> str:
    return _cell_text_normalized(text).replace("：", ":").strip()


_MONTH_HEADER_RE = re.compile(r"^(\d+)\s*个月")


def _month_bucket_n(cell_value: object) -> Optional[int]:
    """从表头如「1个月（26.3月）」提取 n；「1年以上」返回 13。"""
    t = _cell_text_normalized(cell_value)
    if not t:
        return None
    m = _MONTH_HEADER_RE.match(t)
    if m:
        return int(m.group(1))
    if "年以上" in t:
        return 13
    return None


def _find_sheet1_header_row(ws: Worksheet) -> Optional[int]:
    """定位含「主体」「协议账期」的数据表头行（通常为第 2 行）。"""
    max_r = min(ws.max_row or 0, 30)
    max_c = min(ws.max_column or 0, 120)
    best_r: Optional[int] = None
    best_score = 0
    for r in range(1, max_r + 1):
        hits = 0
        has_subject = False
        has_agreement = False
        for c in range(1, max_c + 1):
            h = _norm_h(ws.cell(row=r, column=c).value)
            if h == "主体":
                has_subject = True
                hits += 1
            elif h == "协议账期":
                has_agreement = True
                hits += 1
            elif h == "截止":
                hits += 1
        if has_subject and has_agreement and hits >= best_score:
            best_score = hits
            best_r = r
    return best_r


def _is_invoice_balance_header(text: str) -> bool:
    t = _norm_h(text)
    if t == "当月已开票余额":
        return True
    if "已开票" in t and "余额" in t:
        return True
    if re.match(r"^\d{6,8}.*余额$", t):
        return True
    return False


def _find_col_by_exact(ws: Worksheet, header_row: int, name: str) -> Optional[int]:
    for c in range(1, (ws.max_column or 0) + 1):
        if _norm_h(ws.cell(row=header_row, column=c).value) == name:
            return c
    return None


def _find_col_header_contains(ws: Worksheet, header_row: int, needle: str) -> Optional[int]:
    """表头文本包含 needle 的首列（用于「付款优先级（…）」等长表头）。"""
    for c in range(1, (ws.max_column or 0) + 1):
        h = _cell_text_normalized(ws.cell(row=header_row, column=c).value)
        if needle in h:
            return c
    return None


def _find_payment_plan_col(ws: Worksheet, header_row: int) -> Optional[int]:
    """「4月付款计划」等：表头含「付款计划」且不含「确认」（排除「付款计划确认」）。"""
    for c in range(1, (ws.max_column or 0) + 1):
        h = _cell_text_normalized(ws.cell(row=header_row, column=c).value)
        if "付款计划" in h and "确认" not in h:
            return c
    return None


def _sum_payment_plan_non_empty_numeric(
    ws: Worksheet, header_row: int, c_payment_plan: int, max_r: int
) -> float:
    """表头行以下：付款计划列非空且可解析为数字的单元格求和。"""
    total = 0.0
    for r in range(header_row + 1, max_r + 1):
        cell = ws.cell(row=r, column=c_payment_plan)
        if isinstance(cell, MergedCell):
            continue
        v = cell.value
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        n = _coerce_numeric_for_sum(v)
        if n is not None:
            total += float(n)
    return total


def _sum_payment_plan_by_priority(
    ws: Worksheet,
    header_row: int,
    c_payment_plan: int,
    c_priority: int,
    max_r: int,
    target_pl: int,
) -> float:
    """付款计划列：仅付款优先级 = target_pl 的数据行数值求和（排款写回后统计卡片用）。"""
    total = 0.0
    for r in range(header_row + 1, max_r + 1):
        if _priority_level_0_2(ws.cell(row=r, column=c_priority).value) != target_pl:
            continue
        cell = ws.cell(row=r, column=c_payment_plan)
        if isinstance(cell, MergedCell):
            continue
        n = _coerce_numeric_for_sum(cell.value)
        if n is not None:
            total += float(n)
    return total


def _find_invoice_balance_col(ws: Worksheet, header_row: int) -> Optional[int]:
    for c in range(1, (ws.max_column or 0) + 1):
        raw = ws.cell(row=header_row, column=c).value
        if _is_invoice_balance_header(_cell_text_normalized(raw)):
            return c
    return None


def _find_cutoff_col(ws: Worksheet, header_row: int) -> Optional[int]:
    return _find_col_by_exact(ws, header_row, "截止")


def _find_first_payment_col(ws: Worksheet, header_row: int) -> Optional[int]:
    for c in range(1, (ws.max_column or 0) + 1):
        h = _norm_h(ws.cell(row=header_row, column=c).value)
        if h == "第1笔" or h.startswith("第1笔"):
            return c
    return None


def _find_total_scheduled_col(ws: Worksheet, header_row: int) -> Optional[int]:
    for c in range(1, (ws.max_column or 0) + 1):
        h = _norm_h(ws.cell(row=header_row, column=c).value)
        if h == "合计已排款":
            return c
    return None


def _find_sheet2_category_header_cols(
    ws: Worksheet,
) -> Optional[Tuple[int, int, int, Optional[int]]]:
    """定位 Sheet2 上「类别」「类别优先级」及可选「序号」所在表头行及列号（1-based）。"""
    max_r = min(ws.max_row or 0, 15)
    max_c = min(ws.max_column or 0, 80)
    for r in range(1, max_r + 1):
        c_cat: Optional[int] = None
        c_pri: Optional[int] = None
        c_seq: Optional[int] = None
        for c in range(1, max_c + 1):
            h = _norm_h(ws.cell(row=r, column=c).value)
            if h == "类别":
                c_cat = c
            elif h == "类别优先级":
                c_pri = c
            elif h == "序号":
                c_seq = c
        if c_cat is not None and c_pri is not None:
            return r, c_cat, c_pri, c_seq
    return None


def _load_category_priority_map(
    ws: Worksheet,
    header_row: int,
    c_cat: int,
    c_pri: int,
    c_sheet2_seq: Optional[int],
) -> Dict[str, Tuple[float, int]]:
    """Sheet2：类别 →（类别优先级, Sheet2 序号）；同一类别多行取首次出现行；无「序号」列时序号为 0。"""
    m: Dict[str, Tuple[float, int]] = {}
    for r in range(header_row + 1, (ws.max_row or 0) + 1):
        k = _cell_text_normalized(ws.cell(row=r, column=c_cat).value)
        if not k:
            continue
        if k in m:
            continue
        pv = _coerce_numeric_for_sum(ws.cell(row=r, column=c_pri).value)
        if pv is None:
            continue
        seq_i = 0
        if c_sheet2_seq is not None:
            sv = _coerce_numeric_for_sum(ws.cell(row=r, column=c_sheet2_seq).value)
            if sv is not None:
                seq_i = int(round(float(sv)))
        m[k] = (float(pv), seq_i)
    return m


def _sheet1_serial_sort_key(ws: Worksheet, r: int, c_serial: Optional[int]) -> int:
    if c_serial is None:
        return r
    v = ws.cell(row=r, column=c_serial).value
    if v is None:
        return r
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return r


def _order_candidates_by_category_within_priority(
    ws: Worksheet,
    base: List[Tuple[int, int]],
    c_category: int,
    c_serial: Optional[int],
    cat_map: Dict[str, Tuple[float, int]],
) -> List[Tuple[int, int]]:
    """
    保持付款优先级 0→1→2 分段；每段内：Sheet2 类别优先级 → Sheet2 序号 → Sheet1 序号 → 行号；
    类别为空或不在 Sheet2 的行排在同段末尾，按 Sheet1 序号 → 行号稳定排序。
    """
    by_pl: Dict[int, List[Tuple[int, int, Optional[float], int, int]]] = defaultdict(
        list
    )
    for r, pl in base:
        cat_k = _cell_text_normalized(ws.cell(row=r, column=c_category).value)
        entry = cat_map.get(cat_k) if cat_k else None
        pcat: Optional[float] = None
        pseq2 = 0
        if entry is not None:
            pcat, pseq2 = entry[0], entry[1]
        ser = _sheet1_serial_sort_key(ws, r, c_serial)
        by_pl[pl].append((r, pl, pcat, pseq2, ser))
    out: List[Tuple[int, int]] = []
    for pl in (0, 1, 2):
        bucket = by_pl.get(pl, [])
        matched = [t for t in bucket if t[2] is not None]
        unmatched = [t for t in bucket if t[2] is None]
        matched.sort(key=lambda t: (t[2], t[3], t[4], t[0]))
        unmatched.sort(key=lambda t: (t[4], t[0]))
        for t in matched + unmatched:
            out.append((t[0], t[1]))
    return out


def _find_sheet3_bank_priority_header_cols(ws: Worksheet) -> Optional[Tuple[int, int, int]]:
    """定位 Sheet3 上「银行名称」与「应付款优先级」或「付款优先级」所在表头行及列号（1-based）。"""
    max_r = min(ws.max_row or 0, 20)
    max_c = min(ws.max_column or 0, 40)
    for r in range(1, max_r + 1):
        c_bank: Optional[int] = None
        c_pri: Optional[int] = None
        for c in range(1, max_c + 1):
            h = _norm_h(ws.cell(row=r, column=c).value)
            if h == "银行名称":
                c_bank = c
            elif h == "应付款优先级" or h == "付款优先级":
                if c_pri is None:
                    c_pri = c
        if c_bank is not None and c_pri is not None:
            return r, c_bank, c_pri
    return None


def _load_sheet3_bank_priority_map(
    ws: Worksheet, header_row: int, c_bank: int, c_pri: int
) -> Dict[str, float]:
    """Sheet3：银行名称 → 优先级数值（越小越优先）；同名多行取首次。"""
    m: Dict[str, float] = {}
    for r in range(header_row + 1, (ws.max_row or 0) + 1):
        k = _cell_text_normalized(ws.cell(row=r, column=c_bank).value)
        if not k or k in m:
            continue
        pv = _coerce_numeric_for_sum(ws.cell(row=r, column=c_pri).value)
        if pv is None:
            continue
        m[k] = float(pv)
    return m


def _payment_cols_sorted_by_sheet3_bank_priority(
    ws: Worksheet,
    bank_cols: List[int],
    bank_pri_map: Dict[str, float],
) -> List[int]:
    """
    仅调整内存中的列遍历顺序，不移动工作表上的列。
    已配置银行按 Sheet3 优先级数值升序；未配置视为最低档，排在最后（同档按列号升序）。
    列与 Sheet3 的对应关系：Sheet1 第 3 行该列的银行名称与 Sheet3「银行名称」一致（规范化后匹配）。
    """
    def sort_key(c: int) -> Tuple[int, float, int]:
        nm = _cell_text_normalized(ws.cell(row=3, column=c).value)
        if nm and nm in bank_pri_map:
            return (0, bank_pri_map[nm], c)
        return (1, 0.0, c)

    return sorted(bank_cols, key=sort_key)


def _collect_invoice_cutoff_sum_terms(
    ws: Worksheet,
    header_row: int,
    r: int,
    c_invoice: int,
    c_cutoff: int,
    min_month_threshold: float,
    *,
    non_negative_only: bool = True,
) -> List[Tuple[str, float]]:
    """
    步骤 2～3：对「当月已开票余额」列之后～「截止」列之前各列，读表头得 nMonth；
    仅当 nMonth 可解析且 nMonth ≥ min_month_threshold 时，将本行该列数值计入和（截止列及右侧不参与）。
    默认仅计入数值 ≥ 0 的格（负数跳过）；nPayablesSum 与卡片 1 应付汇总均用此规则。
    """
    terms: List[Tuple[str, float]] = []
    for c in range(c_invoice + 1, c_cutoff):
        mn = _month_bucket_n(ws.cell(row=header_row, column=c).value)
        if mn is None:
            continue
        if float(mn) < min_month_threshold:
            continue
        num = _coerce_numeric_for_sum(ws.cell(row=r, column=c).value)
        if num is None:
            continue
        fv = float(num)
        if non_negative_only and fv < 0.0:
            continue
        terms.append((f"{get_column_letter(c)}{r}", fv))
    return terms


def _format_invoice_cutoff_sum_expr(terms: List[Tuple[str, float]]) -> str:
    """将参与求和的格格式化为 E12+F12+G12=300 或 E12=100。"""
    if not terms:
        return "（无参与求和的数值格）=0"
    addrs = "+".join(addr for addr, _ in terms)
    total = sum(v for _, v in terms)
    return f"{addrs}={total:g}"


def _sum_over_invoice_cutoff_window(
    ws: Worksheet,
    header_row: int,
    r: int,
    c_invoice: int,
    c_cutoff: int,
    min_month_threshold: float,
    *,
    non_negative_only: bool = True,
) -> float:
    terms = _collect_invoice_cutoff_sum_terms(
        ws,
        header_row,
        r,
        c_invoice,
        c_cutoff,
        min_month_threshold,
        non_negative_only=non_negative_only,
    )
    return sum(v for _, v in terms)


_LIGHT_GREEN_FILL = PatternFill(
    patternType="solid", fgColor="FFC6EFCE", bgColor="FFC6EFCE"
)


def _gradient_left_base_color(cell: Union[Cell, MergedCell]) -> Color:
    """渐变左侧：沿用写入前单元格原有填充色（无填充或无法解析时退化为白）。"""
    if isinstance(cell, MergedCell):
        return Color("FFFFFFFF")
    f = cell.fill
    pt = getattr(f, "patternType", None)
    if pt in (None, "none"):
        return Color("FFFFFFFF")
    if pt == "solid":
        col = getattr(f, "fgColor", None)
        if col is not None:
            rgb = getattr(col, "rgb", None)
            if rgb:
                s = str(rgb).strip().upper()
                if len(s) == 6 and all(c in "0123456789ABCDEF" for c in s):
                    return Color("FF" + s)
                if len(s) == 8 and all(c in "0123456789ABCDEF" for c in s):
                    return Color(s)
    return Color("FFFFFFFF")


def _fill_partial_light_green_from_right(
    ws: Worksheet, row: int, col: int, fraction_green_from_right: float
) -> None:
    """
    仅本格：把单元格整体宽度视为 100%，右侧 fraction_green_from_right（0~1）为浅绿，
    左侧 (1−该比例) 保持原格填充色（从写入渐变前的 cell.fill 读取）。
    fraction =（锚点格数值 − nPayablesSum×(nPeriodPercent−阈值百分数)/100）/ 锚点格数值。
    线性渐变 degree=0：位置 0=左、1=右；分界在 1−p，过渡带极窄以免吃掉比例。
    """
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    p = max(0.0, min(1.0, float(fraction_green_from_right)))
    if p <= 0.0:
        return
    if p >= 1.0 - 1e-12:
        cell.fill = copy(_LIGHT_GREEN_FILL)
        return

    left_end = 1.0 - p
    base = _gradient_left_base_color(cell)
    # 过渡带不得超过较窄一侧的 2%，且上限极小，保证「约 p 宽为绿、约 1−p 为原底色」
    span = min(p, left_end)
    eps = min(5e-5, max(1e-9, span * 0.02))
    if left_end <= eps * 2:
        cell.fill = copy(_LIGHT_GREEN_FILL)
        return

    t0 = 0.0
    t1 = max(t0 + 1e-12, left_end - eps)
    t2 = min(1.0 - 1e-12, left_end + eps)
    t3 = 1.0
    if t2 <= t1 + 1e-12:
        t2 = min(t3, t1 + 2e-12)
    if t2 >= t3:
        cell.fill = copy(_LIGHT_GREEN_FILL)
        return

    green = Color("FFC6EFCE")
    stops = (Stop(base, t0), Stop(base, t1), Stop(green, t2), Stop(green, t3))
    cell.fill = GradientFill(type="linear", degree=0, stop=stops)


def _cell_is_numeric_for_green_highlight(value: object) -> bool:
    """非空、且为有限数值型（排除 bool）、且大于 0 才标浅绿；其余不处理。"""
    if value is None or isinstance(value, bool):
        return False
    if isinstance(value, str):
        return False
    if isinstance(value, (datetime, date)):
        return False
    if isinstance(value, (int, float)):
        if isinstance(value, float) and not math.isfinite(value):
            return False
        return float(value) > 0.0
    return False


def _set_cell_light_green_if_numeric(ws: Worksheet, row: int, col: int) -> None:
    """
    与 nPayablesSum / 锚点判定一致：用 _coerce_numeric_for_sum 解析（含文本数字），
    仅当解析为有限数且 >0 时标浅绿（避免纯 int/float 判断导致「数字为文本」时整段不着色）。
    """
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    n = _coerce_numeric_for_sum(cell.value)
    if n is None:
        return
    fv = float(n)
    if not math.isfinite(fv) or fv <= 0.0:
        return
    cell.fill = copy(_LIGHT_GREEN_FILL)


def _cols_qualifying_month(
    ws: Worksheet, header_row: int, r: int, c_invoice: int, c_cutoff: int, n_period: float
) -> List[int]:
    out: List[int] = []
    for c in range(c_invoice + 1, c_cutoff):
        mn = _month_bucket_n(ws.cell(row=header_row, column=c).value)
        if mn is None:
            continue
        if float(mn) < float(n_period):
            continue
        out.append(c)
    return out


def _find_last_nonempty_numeric_col(
    ws: Worksheet, r: int, cols: List[int],
) -> Optional[int]:
    """自右向左最后一个可解析且数值 ≥ 0 的列（与 nPayablesSum 不计负数一致）。"""
    for c in reversed(cols):
        n = _coerce_numeric_for_sum(ws.cell(row=r, column=c).value)
        if n is not None and float(n) >= 0.0:
            return c
    return None


def _find_first_col_by_month_in_window(
    ws: Worksheet, header_row: int, c_invoice: int, c_cutoff: int, target_mn: int
) -> Optional[int]:
    for c in range(c_invoice + 1, c_cutoff):
        mn = _month_bucket_n(ws.cell(row=header_row, column=c).value)
        if mn == target_mn:
            return c
    return None


def _sum_numeric_from_col_right_in_window(
    ws: Worksheet, r: int, c_start: int, c_cutoff: int
) -> float:
    """自 c_start 至截止列前求和；仅计入数值 ≥ 0 的格（与 nPayablesSum 一致）。"""
    s = 0.0
    for c in range(c_start, c_cutoff):
        n = _coerce_numeric_for_sum(ws.cell(row=r, column=c).value)
        if n is not None:
            fv = float(n)
            if fv >= 0.0:
                s += fv
    return s


def _highlight_equal_band(
    ws: Worksheet, r: int, c_anchor: int, c_cutoff: int,
) -> None:
    for c in range(c_anchor, c_cutoff):
        _set_cell_light_green_if_numeric(ws, r, c)


def _highlight_priority0_month_window(
    ws: Worksheet,
    header_row: int,
    r: int,
    c_invoice: int,
    c_cutoff: int,
    n_period: float,
) -> None:
    """
    优先级 0：表头 mn≥nPeriodCount 的列（「当月已开票余额」与「截止」之间），
    本行单元格解析为有限数且 >0 则整格浅绿（截止列及其右侧不参与）。
    """
    qual = _cols_qualifying_month(ws, header_row, r, c_invoice, c_cutoff, n_period)
    for c in qual:
        _set_cell_light_green_if_numeric(ws, r, c)


def _normalize_anchor_cell_strip_slash_suffix(ws: Worksheet, r: int, c: int) -> Optional[float]:
    """
    锚点格若为「金额/…」文本，去掉斜杠及后缀，写回左侧金额（数值），并返回该金额。
    已是纯数值且 >0 则不改写，直接返回。
    """
    cell = ws.cell(row=r, column=c)
    if isinstance(cell, MergedCell):
        return None
    v = cell.value
    if isinstance(v, str) and "/" in v:
        left = v.split("/", 1)[0].strip()
        n = _coerce_numeric_for_sum(left)
        if n is None or float(n) <= 0:
            return None
        fv = float(n)
        if abs(fv - round(fv)) < 1e-9:
            cell.value = int(round(fv))
        else:
            cell.value = fv
        return fv
    if _cell_is_numeric_for_green_highlight(v):
        return float(v)
    return None


def _highlight_greater_band(
    ws: Worksheet,
    header_row: int,
    r: int,
    c_invoice: int,
    c_cutoff: int,
    current_mn: int,
    current_c: int,
    npct: float,
    thr_pct: float,
    n_payables_sum: float,
) -> None:
    c_plus = _find_first_col_by_month_in_window(
        ws, header_row, c_invoice, c_cutoff, current_mn + 1
    )
    if c_plus is not None:
        for c in range(c_plus, c_cutoff):
            _set_cell_light_green_if_numeric(ws, r, c)
    fv = _normalize_anchor_cell_strip_slash_suffix(ws, r, current_c)
    if fv is None:
        return
    # 百分数差先 /100 再乘 nPayablesSum，与循环内 nPeriodPercent 分母一致
    delta_amt = float(n_payables_sum) * (float(npct) - float(thr_pct)) / 100.0
    suffix = fv - delta_amt
    if suffix < 0:
        suffix = 0.0
    # 整格宽度=100%：右侧 prop_green 为浅绿，左侧 1−prop_green 保持原格填充色（与分式一致）
    prop_green = (suffix / fv) if fv > 0 else 0.0
    _fill_partial_light_green_from_right(ws, r, current_c, prop_green)


def _compute_ntotal_priority12_with_highlights(
    ws: Worksheet,
    header_row: int,
    r: int,
    c_invoice: int,
    c_cutoff: int,
    n_period: float,
    thr_pct: float,
) -> float:
    """
    优先级 1 或 2 行：nTotalSum = nPayablesSum × (thr_pct/100)（thr_pct 即界面 nPercentOne 或 nPercentTwo）。

    nPayablesSum 定义同 _sum_over_invoice_cutoff_window。锚点：mn≥nPeriodCount 的遍历区间内，
    自右向左最后一个可解析为数值的非空列；nPeriodPercent = (该列单元格数值 / nPayablesSum) × 100。

    - 若 nPeriodPercent == thr_pct：从该列起至截止列前，遍历范围内数值非空且>0 的格整格浅绿（_highlight_equal_band）。
    - 若 nPeriodPercent < thr_pct：按表头月份减 1 向左换列，nPeriodPercent 改为（该列及右侧数值之和
      / nPayablesSum）×100，循环直至 nPeriodPercent≥thr_pct 或无法再降月。
      若最终等于阈值则整段浅绿；若大于阈值则走「大于」分支。
    - 若 nPeriodPercent > thr_pct：（nMonth+1）月列及其右侧遍历范围内数值非空且>0 的格整格浅绿；
      当前 nMonth 列按 (单元格数值 − nPayablesSum×(nPeriodPercent−thr_pct)/100) / 单元格数值 作为
      从右向左的浅绿占比，做分格渐变（_highlight_greater_band / _fill_partial_light_green_from_right）。

    约定：锚点列上初始 nPeriodPercent 不大于阈值的典型数据；若数据导致首算即大于阈值，仍按「大于」分支着色。
    """
    n_payables = _sum_over_invoice_cutoff_window(
        ws, header_row, r, c_invoice, c_cutoff, n_period
    )
    n_total_sum = float(n_payables) * float(thr_pct) / 100.0

    qual = _cols_qualifying_month(ws, header_row, r, c_invoice, c_cutoff, n_period)
    c_anchor = _find_last_nonempty_numeric_col(ws, r, qual)
    if c_anchor is None or n_payables <= 0:
        return n_total_sum

    mn0 = _month_bucket_n(ws.cell(row=header_row, column=c_anchor).value)
    if mn0 is None:
        return n_total_sum

    v0 = _coerce_numeric_for_sum(ws.cell(row=r, column=c_anchor).value)
    if v0 is None or float(v0) < 0.0:
        return n_total_sum

    npct = (float(v0) / float(n_payables)) * 100.0
    thr = float(thr_pct)

    if math.isclose(npct, thr, rel_tol=0.0, abs_tol=1e-6):
        _highlight_equal_band(ws, r, c_anchor, c_cutoff)
        return n_total_sum

    if npct > thr + 1e-9:
        _highlight_greater_band(
            ws,
            header_row,
            r,
            c_invoice,
            c_cutoff,
            int(mn0),
            c_anchor,
            npct,
            thr,
            float(n_payables),
        )
        return n_total_sum

    if npct < thr - 1e-9:
        current_c = c_anchor
        current_mn = int(mn0)
        max_steps = 200
        for _ in range(max_steps):
            c_col = _find_first_col_by_month_in_window(
                ws, header_row, c_invoice, c_cutoff, current_mn - 1
            )
            if c_col is None:
                break
            current_mn -= 1
            current_c = c_col
            total = _sum_numeric_from_col_right_in_window(ws, r, current_c, c_cutoff)
            npct = (total / float(n_payables)) * 100.0
            if not (npct < thr - 1e-9):
                break

        if math.isclose(npct, thr, rel_tol=0.0, abs_tol=1e-6):
            _highlight_equal_band(ws, r, current_c, c_cutoff)
        elif npct > thr + 1e-9:
            _highlight_greater_band(
                ws,
                header_row,
                r,
                c_invoice,
                c_cutoff,
                current_mn,
                current_c,
                npct,
                thr,
                float(n_payables),
            )

    return n_total_sum


def _compute_ntotal_for_priority(
    ws: Worksheet,
    header_row: int,
    r: int,
    c_invoice: int,
    c_cutoff: int,
    n_period: float,
    priority_level: int,
    priority1_plan_pct: float,
    priority2_plan_pct: float,
) -> float:
    """
    计算本行 nTotalSum（写入「笔」列与付款计划、扣减排款余额用）。

    - 优先级 0：nTotalSum = nPayablesSum；并对 mn≥nPeriodCount 至截止列前、本行数值非空且>0 的格标浅绿。
    - 优先级 1：nTotalSum = nPayablesSum × (priority1_plan_pct/100)；着色见 _compute_ntotal_priority12_with_highlights（nPercentOne）。
    - 优先级 2：nTotalSum = nPayablesSum × (priority2_plan_pct/100)；同上（nPercentTwo）。
    """
    if priority_level == 0:
        n_payables = _sum_over_invoice_cutoff_window(
            ws, header_row, r, c_invoice, c_cutoff, n_period
        )
        _highlight_priority0_month_window(
            ws, header_row, r, c_invoice, c_cutoff, n_period
        )
        return n_payables

    thr = priority1_plan_pct if priority_level == 1 else priority2_plan_pct
    return _compute_ntotal_priority12_with_highlights(
        ws, header_row, r, c_invoice, c_cutoff, n_period, thr
    )


def _protocol_period_months(value: object) -> Optional[float]:
    """协议账期 ÷ 30 → nPeriodCount；无法解析则 None。"""
    if value is None:
        return None
    s = str(value).strip()
    if not s or s in ("-", "—", "－"):
        return None
    n = _coerce_numeric_for_sum(value)
    if n is None or n <= 0:
        return None
    return n / 30.0


def summarize_card1_payable_from_total_sheet(path: str) -> Optional[Card1PayableTotals]:
    """
    从 TotalSheet 总表 Sheet1 汇总卡片 1 应付额（仅 .xlsx）。与下列步骤一致：

    1. 数据区间：表头行定位后，列范围为「当月已开票余额」列之后至「截止」列之前；截止列及右侧不参与。
    2. 表头各列解析月份数 nMonth（「N个月」取 N；「年以上」等见 _month_bucket_n）。
    3. 行内 nPeriodCount = 协议账期÷30；仅保留 nMonth ≥ nPeriodCount 的列参与本行求和。
    4. 付款优先级=0 的行：上述列内数值 ≥0 的单元格求和，累加为 nCount0。
    5. 付款优先级=1 的行：同上累加为 nCount1。
    6. 付款优先级=2 的行：同上累加为 nCount2。
    （与智能排款 nPayablesSum 相同：负数不计。）
    7. nCountTotal = nCount0 + nCount1 + nCount2。
    8. 供界面展示四行；展示后由 main 将内存变量 nCount0/1/2/Total 置 0（界面数字不变）。

    行筛选：非空主体且付款优先级为 0/1/2；协议账期无法解析则跳过该行。
    步骤 4 对优先级 0 逐行打印求和明细到控制台（如 Z17+AA17=138400）。
    """
    if _is_xls(path):
        return None
    wb = None
    try:
        wb = load_workbook(path, read_only=False, data_only=True)
        ws = wb.worksheets[0]
        header_row = _find_sheet1_header_row(ws)
        if header_row is None:
            return None
        c_subject = _find_col_by_exact(ws, header_row, "主体")
        c_agreement = _find_col_by_exact(ws, header_row, "协议账期")
        c_invoice = _find_invoice_balance_col(ws, header_row)
        c_cutoff = _find_cutoff_col(ws, header_row)
        c_priority = _find_col_header_contains(ws, header_row, "付款优先级")
        if (
            c_subject is None
            or c_agreement is None
            or c_invoice is None
            or c_cutoff is None
            or c_priority is None
        ):
            return None
        if c_invoice >= c_cutoff:
            return None
        max_r = ws.max_row or 0
        n0 = 0.0
        n1 = 0.0
        n2 = 0.0
        # logged_n0_step4 = False
        for r in range(header_row + 1, max_r + 1):
            pl = _priority_level_0_2(ws.cell(row=r, column=c_priority).value)
            if pl is None:
                continue
            subj = _cell_text_normalized(ws.cell(row=r, column=c_subject).value)
            if not subj:
                continue
            n_period = _protocol_period_months(ws.cell(row=r, column=c_agreement).value)
            if n_period is None:
                continue
            if pl == 0:
                terms = _collect_invoice_cutoff_sum_terms(
                    ws,
                    header_row,
                    r,
                    c_invoice,
                    c_cutoff,
                    n_period,
                )
                row_sum = float(sum(v for _, v in terms))
                # if not logged_n0_step4:
                #     print(
                #         "[卡片1应付·步骤4] 筛选付款优先级=0，"
                #         "nMonth≥nPeriodCount 且单元格数值≥0 的列求和累计 nCount0",
                #         flush=True,
                #     )
                #     logged_n0_step4 = True
                n0 += row_sum
                # print(
                #     f"[卡片1应付·步骤4] 第{r}行 主体={subj} "
                #     f"nPeriodCount={n_period:g} | "
                #     f"{_format_invoice_cutoff_sum_expr(terms)} | "
                #     f"本行={row_sum:,.2f} | 累计nCount0={n0:,.2f}",
                #     flush=True,
                # )
            else:
                row_sum = float(
                    _sum_over_invoice_cutoff_window(
                        ws,
                        header_row,
                        r,
                        c_invoice,
                        c_cutoff,
                        n_period,
                    )
                )
                if pl == 1:
                    n1 += row_sum
                elif pl == 2:
                    n2 += row_sum
        # if logged_n0_step4:
        #     print(
        #         f"[卡片1应付·步骤4] nCount0 合计={n0:,.2f}",
        #         flush=True,
        #     )
        nt = n0 + n1 + n2
        return Card1PayableTotals(
            n_count0=n0, n_count1=n1, n_count2=n2, n_count_total=nt
        )
    except Exception:
        return None
    finally:
        if wb is not None:
            wb.close()


def _split_tokens(s: str) -> List[str]:
    parts: List[str] = []
    for chunk in re.split(r"[/、，,|；;]+", s):
        t = chunk.strip()
        if t:
            parts.append(t)
    return parts


def _subject_contained_in_pay_region(row_subject: str, pay_cell: object) -> bool:
    """条件 2a：当前行主体包含于第七行「可支付主体」单元格全文即可。"""
    sub = row_subject.strip()
    if not sub:
        return False
    cell = _cell_text_normalized(pay_cell)
    if not cell:
        return False
    return sub in cell


def _payment_method_any_token_match(row_method: str, col_method_cell: object) -> bool:
    """条件 2b：当前行收款方式拆分后，与第八行该列收款方式拆分结果任一相同或包含关系即通过。"""
    rm = row_method.strip()
    if not rm:
        return False
    cm = _cell_text_normalized(col_method_cell)
    if not cm:
        return False
    row_toks = _split_tokens(rm)
    if not row_toks:
        row_toks = [rm]
    col_toks = _split_tokens(cm)
    if not col_toks:
        col_toks = [cm]
    for r in row_toks:
        if not r:
            continue
        if r in cm:
            return True
        for c in col_toks:
            if not c:
                continue
            if r == c or r in c or c in r:
                return True
    return False


def _read_schedule_balance(ws: Worksheet, col: int) -> float:
    """
    第 5 行「排款余额」；若为空则用第 4 行「排款总额」作为初始可排额度
    （兼容尚未维护排款余额公式的模版）。
    """
    v5 = ws.cell(row=5, column=col).value
    n5 = _coerce_numeric_for_sum(v5)
    if n5 is not None:
        return float(n5)
    v4 = ws.cell(row=4, column=col).value
    n4 = _coerce_numeric_for_sum(v4)
    if n4 is not None:
        return float(n4)
    return 0.0


def _parse_cell_datetime(value: object) -> Optional[datetime]:
    """将单元格解析为 datetime（用于放款日、付款截止日比较）。"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date) and not isinstance(value, datetime):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            return from_excel(float(value))
        except (TypeError, ValueError, OSError):
            return None
    s = str(value).strip()
    if not s or s in ("-", "—", "－"):
        return None
    if len(s) >= 19:
        try:
            return datetime.strptime(s[:19], "%Y-%m-%d %H:%M:%S")
        except ValueError:
            pass
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt)
        except ValueError:
            continue
    return None


def _loan_deadline_ok(ws: Worksheet, pay_col: int, deadline_cell: object) -> bool:
    """条件 3：第 6 行放款日期 ≤ 当前行付款截止日期（同日可）。"""
    loan_raw = ws.cell(row=6, column=pay_col).value
    loan_dt = _parse_cell_datetime(loan_raw)
    due_dt = _parse_cell_datetime(deadline_cell)
    if loan_dt is None or due_dt is None:
        return False
    return loan_dt <= due_dt


def _priority_level_0_2(value: object) -> Optional[int]:
    """付款优先级为 0、1、2 时返回对应整数，否则 None。"""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        try:
            x = float(value)
        except (TypeError, ValueError):
            return None
        if x == 0.0:
            return 0
        if x == 1.0:
            return 1
        if x == 2.0:
            return 2
        return None
    s = str(value).strip()
    if not s:
        return None
    try:
        x = float(s)
    except ValueError:
        return None
    if x == 0.0:
        return 0
    if x == 1.0:
        return 1
    if x == 2.0:
        return 2
    return None


def _find_row4_payout_gross_col(ws: Worksheet) -> Optional[int]:
    for c in range(1, (ws.max_column or 0) + 1):
        t = _cell_text_normalized(ws.cell(row=4, column=c).value)
        if _norm_header_label(t) == "排款总额":
            return c
    return None


def _find_row1_schedule_plan_column_span(ws: Worksheet) -> Optional[Tuple[int, int]]:
    max_c = ws.max_column or 0
    if max_c <= 0:
        return None
    for mr in ws.merged_cells.ranges:
        if mr.min_row > 1 or mr.max_row < 1:
            continue
        tl = _cell_text_normalized(ws.cell(row=1, column=mr.min_col).value)
        if "排款计划" in tl:
            return (mr.min_col, mr.max_col)
    for c in range(1, max_c + 1):
        tl = _cell_text_normalized(ws.cell(row=1, column=c).value)
        if "排款计划" not in tl:
            continue
        for mr in ws.merged_cells.ranges:
            if mr.min_row <= 1 <= mr.max_row and mr.min_col <= c <= mr.max_col:
                return (mr.min_col, mr.max_col)
        return (c, c)
    return None


def _clear_cell_if_writable(ws: Worksheet, row: int, col: int) -> None:
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = None


def _copy_cell_as_is(src: Cell, dst: Cell) -> None:
    if isinstance(dst, MergedCell):
        return
    dst.value = src.value
    if src.has_style:
        dst.number_format = src.number_format
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)


def _prepare_sheet_three_steps(
    ws: Worksheet,
    header_row: int,
    c_payment_plan: int,
    c_total_sched: int,
) -> None:
    """智能排款前：第 9 行起清空「付款计划」～「合计已排款」矩形区；清空第 1 行「排款计划」列区第 9 行起；第 4 行横向区复制到第 5 行。
    nPayablesSum 为行内重算，无跨行缓存；清空填写区后不从旧格累加。"""
    max_r = ws.max_row or 0
    if max_r <= 0:
        return

    c_blk_lo = min(c_payment_plan, c_total_sched)
    c_blk_hi = max(c_payment_plan, c_total_sched)
    if max_r >= 9:
        for r in range(9, max_r + 1):
            for c in range(c_blk_lo, c_blk_hi + 1):
                _clear_cell_if_writable(ws, r, c)

    span = _find_row1_schedule_plan_column_span(ws)
    if span is not None:
        c_lo, c_hi = span
        for r in range(9, max_r + 1):
            for c in range(c_lo, c_hi + 1):
                _clear_cell_if_writable(ws, r, c)

    c_gross = _find_row4_payout_gross_col(ws)
    if c_gross is None:
        return
    start_c = c_gross + 1
    end_c = c_total_sched
    if start_c > end_c:
        return
    for c in range(start_c, end_c + 1):
        src = ws.cell(row=4, column=c)
        if isinstance(src, MergedCell):
            continue
        dst = ws.cell(row=5, column=c)
        _copy_cell_as_is(src, dst)


def run_smart_schedule_on_total_sheet(
    path: str,
    priority1_plan_pct: float,
    priority2_plan_pct: float,
) -> SmartScheduleResult:
    """
    在排款计划总表 Sheet1 上执行智能排款（仅 .xlsx）。

    与界面「第一/二优先级支付计划占应付额」整数百分比（0～100）对应：程序内以
    priority1_plan_pct / priority2_plan_pct 作为 nPercentOne / nPercentTwo（除以 100 后乘 nPayablesSum）。

    每行 nTotalSum 与着色由 _compute_ntotal_for_priority 实现；写「笔」列与付款计划、扣排款余额后，
    SmartScheduleResult：priority0_total 为所有付款优先级=0 行的 nTotalSum 之和（严格按账期排款额）；
    priority1_total / priority2_total 为表头行以下「付款计划」列中该行「付款优先级」为 1/2 的数值之和；
    界面「本月合计排款总额」= 三者之和（main 中 lbl_total_payable_value）。

    表内排序与列匹配逻辑：表初始化后按优先级 0、1、2 顺序处理各行；同档内按 Sheet2「类别优先级」、
    Sheet2「序号」（表头含「序号」列时）、Sheet1「序号」稳定排序；类别未匹配行在同档末尾按 Sheet1「序号」
    稳定排序；各「笔」列按 Sheet3 银行优先级数值升序遍历。
    """
    if _is_xls(path):
        raise ValueError("智能排款仅支持 .xlsx 格式的排款计划总表。")

    wb = load_workbook(path, read_only=False, data_only=False)
    try:
        ws = wb.worksheets[0]
        header_row = _find_sheet1_header_row(ws)
        if header_row is None:
            raise ValueError("未在 Sheet1 中找到含「主体」「协议账期」的表头行。")

        c_subject = _find_col_by_exact(ws, header_row, "主体")
        c_agreement = _find_col_by_exact(ws, header_row, "协议账期")
        c_method = _find_col_by_exact(ws, header_row, "收款方式")
        c_invoice = _find_invoice_balance_col(ws, header_row)
        c_cutoff = _find_cutoff_col(ws, header_row)
        c_pay0 = _find_first_payment_col(ws, header_row)
        c_total_sched = _find_total_scheduled_col(ws, header_row)
        c_priority = _find_col_header_contains(ws, header_row, "付款优先级")
        c_due = _find_col_by_exact(ws, header_row, "付款截止日期")
        c_payment_plan = _find_payment_plan_col(ws, header_row)
        c_category = _find_col_by_exact(ws, header_row, "类别")
        c_serial = _find_col_by_exact(ws, header_row, "序号")

        missing = []
        if c_subject is None:
            missing.append("主体")
        if c_agreement is None:
            missing.append("协议账期")
        if c_method is None:
            missing.append("收款方式")
        if c_invoice is None:
            missing.append("当月已开票余额（或同义「*余额」列）")
        if c_cutoff is None:
            missing.append("截止")
        if c_pay0 is None or c_total_sched is None:
            missing.append("第1笔 / 合计已排款")
        if c_priority is None:
            missing.append("付款优先级")
        if c_due is None:
            missing.append("付款截止日期")
        if c_payment_plan is None:
            missing.append("付款计划（表头含「付款计划」且非「付款计划确认」）")
        if c_category is None:
            missing.append("类别")
        if missing:
            raise ValueError("表头缺少必要列：" + "、".join(missing))

        if len(wb.worksheets) < 2:
            raise ValueError("排款计划总表第二页（Sheet2）不存在，无法读取「类别」「类别优先级」对照。")
        ws2 = wb.worksheets[1]
        sheet2_cat = _find_sheet2_category_header_cols(ws2)
        if sheet2_cat is None:
            raise ValueError("未在 Sheet2 表头区域找到「类别」与「类别优先级」列。")
        h2_cat, c2_cat_col, c2_pri_col, c2_seq_col = sheet2_cat
        cat_map = _load_category_priority_map(
            ws2, h2_cat, c2_cat_col, c2_pri_col, c2_seq_col
        )

        if len(wb.worksheets) < 3:
            raise ValueError(
                "排款计划总表第三页（Sheet3）不存在，无法读取银行「应付款优先级」或「付款优先级」对照。"
            )
        ws3 = wb.worksheets[2]
        sheet3_hdr = _find_sheet3_bank_priority_header_cols(ws3)
        if sheet3_hdr is None:
            raise ValueError(
                "未在 Sheet3 表头区域找到「银行名称」与「应付款优先级」或「付款优先级」列。"
            )
        h3_bank, c3_bank_col, c3_pri_col = sheet3_hdr
        bank_pri_map = _load_sheet3_bank_priority_map(ws3, h3_bank, c3_bank_col, c3_pri_col)

        assert c_payment_plan is not None and c_total_sched is not None
        _prepare_sheet_three_steps(ws, header_row, c_payment_plan, c_total_sched)

        assert c_invoice is not None and c_cutoff is not None
        if c_invoice >= c_cutoff:
            raise ValueError("「当月已开票余额」列须位于「截止」列左侧。")

        bank_cols: List[int] = list(range(c_pay0, c_total_sched))
        if not bank_cols:
            raise ValueError("「第1笔」须位于「合计已排款」左侧，且两列之间至少有一列可排款。")
        payment_cols: List[int] = _payment_cols_sorted_by_sheet3_bank_priority(
            ws, bank_cols, bank_pri_map
        )

        max_r = ws.max_row or 0
        candidates: List[Tuple[int, int]] = []
        for r in range(header_row + 1, max_r + 1):
            pl = _priority_level_0_2(ws.cell(row=r, column=c_priority).value)
            if pl is None:
                continue
            subj = _cell_text_normalized(ws.cell(row=r, column=c_subject).value)
            if not subj:
                continue
            candidates.append((r, pl))
        candidates = _order_candidates_by_category_within_priority(
            ws, candidates, c_category, c_serial, cat_map
        )

        rows_priority0 = sum(1 for _r, pl in candidates if pl == 0)
        rows_priority1 = sum(1 for _r, pl in candidates if pl == 1)
        rows_priority2 = sum(1 for _r, pl in candidates if pl == 2)
        rows_scanned = 0
        rows_written = 0
        rows_skipped = 0
        priority0_total = 0.0

        bal_cache: Dict[int, float] = {}

        def _col_balance(c: int) -> float:
            if c not in bal_cache:
                bal_cache[c] = _read_schedule_balance(ws, c)
            return bal_cache[c]

        def _reduce_balance(c: int, amount: float) -> None:
            new_b = _col_balance(c) - amount
            bal_cache[c] = new_b
            ws.cell(row=5, column=c, value=new_b)

        for r, prio in candidates:
            subj = _cell_text_normalized(ws.cell(row=r, column=c_subject).value)
            n_period = _protocol_period_months(ws.cell(row=r, column=c_agreement).value)
            if n_period is None:
                rows_skipped += 1
                continue

            # nTotalSum：仅此一处按当前行 priority 计算；下方匹配列写入均使用本变量。
            n_total_sum = _compute_ntotal_for_priority(
                ws,
                header_row,
                r,
                c_invoice,
                c_cutoff,
                n_period,
                prio,
                priority1_plan_pct,
                priority2_plan_pct,
            )

            if prio == 0:
                priority0_total += n_total_sum

            rows_scanned += 1

            if n_total_sum <= 0:
                rows_skipped += 1
                continue

            row_method = _cell_text_normalized(ws.cell(row=r, column=c_method).value)
            due_cell = ws.cell(row=r, column=c_due).value

            placed = False
            for c in payment_cols:
                bal = _col_balance(c)
                if bal < n_total_sum:
                    continue
                if not _subject_contained_in_pay_region(subj, ws.cell(row=7, column=c).value):
                    continue
                if not _payment_method_any_token_match(row_method, ws.cell(row=8, column=c).value):
                    continue
                if not _loan_deadline_ok(ws, c, due_cell):
                    continue
                ws.cell(row=r, column=c, value=n_total_sum)
                _reduce_balance(c, n_total_sum)
                assert c_payment_plan is not None
                old_plan = _coerce_numeric_for_sum(
                    ws.cell(row=r, column=c_payment_plan).value
                )
                plan_base = float(old_plan) if old_plan is not None else 0.0
                ws.cell(row=r, column=c_payment_plan, value=plan_base + n_total_sum)
                rows_written += 1
                placed = True
                break
            if not placed:
                rows_skipped += 1

        plan_sum = _sum_payment_plan_non_empty_numeric(
            ws, header_row, c_payment_plan, max_r
        )
        assert c_priority is not None
        priority1_total = _sum_payment_plan_by_priority(
            ws, header_row, c_payment_plan, c_priority, max_r, 1
        )
        priority2_total = _sum_payment_plan_by_priority(
            ws, header_row, c_payment_plan, c_priority, max_r, 2
        )
        wb.save(path)
    finally:
        wb.close()

    payable = sum_sheet1_row4_after_payout_through_total_scheduled(path)
    reserved: Optional[float] = None
    if payable is not None:
        reserved = float(payable) - plan_sum

    return SmartScheduleResult(
        path=path,
        rows_priority0=rows_priority0,
        rows_priority1=rows_priority1,
        rows_priority2=rows_priority2,
        rows_scanned=rows_scanned,
        rows_written=rows_written,
        rows_skipped=rows_skipped,
        reserved_total=reserved,
        priority0_total=priority0_total,
        priority1_total=priority1_total,
        priority2_total=priority2_total,
    )
