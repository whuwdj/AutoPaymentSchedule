# -*- coding: utf-8 -*-
"""
排款明细表数据生成：总表 Sheet1 为数据源，按明细模板各 Sheet 名与第 3 行银行列匹配，
从第 9 行起取数，按第 2 行表头映射写入模板（智能排款成功后由 main 调用）。
"""

from __future__ import annotations

import unicodedata
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Border, Side
from openpyxl.worksheet.worksheet import Worksheet

from excel_service import (
    _cell_text_normalized,
    _coerce_numeric_for_sum,
    _norm_header_label,
)

ROW_HEADERS = 2
ROW_BANK = 3
ROW_DATA_B = 9
ROW_DATA_A = 3

_THIN_SIDE = Side(style="thin", color="000000")
_DATA_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE
)


def _norm_match_key(s: str) -> str:
    s = unicodedata.normalize("NFKC", s or "")
    s = "".join(c for c in s if not c.isspace())
    return s.casefold()


def _header_map_key(cell_val: object) -> str:
    t = _norm_header_label(_cell_text_normalized(cell_val))
    t = unicodedata.normalize("NFKC", t)
    return "".join(ch for ch in t if not ch.isspace())


def _strip_trailing_sheet_suffix(title: str) -> str:
    t = (title or "").strip()
    for suf in ("排款明细",):
        if t.endswith(suf):
            t = t[: -len(suf)].strip()
    return t


def _sheet_title_matches_bank_cell(sheet_title: str, bank_cell_value: object) -> bool:
    b = _cell_text_normalized(bank_cell_value)
    if not b:
        return False
    bk = _norm_match_key(b)
    if not bk:
        return False
    full = _norm_match_key(sheet_title.strip())
    core = _norm_match_key(_strip_trailing_sheet_suffix(sheet_title))
    return bk == full or bk == core


def _find_col_in_rows(
    ws: Worksheet, want_norm: str, rows_1based: Tuple[int, ...]
) -> Optional[int]:
    max_c = ws.max_column or 0
    for r in rows_1based:
        for c in range(1, max_c + 1):
            t = _norm_header_label(_cell_text_normalized(ws.cell(row=r, column=c).value))
            if t == want_norm:
                return c
    return None


def _find_bank_scan_bounds(ws: Worksheet) -> Tuple[int, int]:
    """第 3 行「银行名称」右侧至「合计已排款」列（表头在第 2 或第 3 行）为横向扫描区间。"""
    max_c = ws.max_column or 0
    bank_lbl_c = _find_col_in_rows(ws, "银行名称", (ROW_BANK,))
    if bank_lbl_c is None:
        bank_lbl_c = 0
    end_c = _find_col_in_rows(ws, "合计已排款", (ROW_BANK, ROW_HEADERS))
    if end_c is None or end_c <= bank_lbl_c:
        end_c = max_c
    start_c = bank_lbl_c + 1 if bank_lbl_c else 1
    return start_c, end_c


def _find_bank_column_for_sheet(ws_b: Worksheet, sheet_title: str) -> Optional[int]:
    start_c, end_c = _find_bank_scan_bounds(ws_b)
    for c in range(start_c, end_c + 1):
        v = ws_b.cell(row=ROW_BANK, column=c).value
        if _sheet_title_matches_bank_cell(sheet_title, v):
            return c
    return None


def _collect_data_rows_bank_numeric(ws_b: Worksheet, c_bank: int) -> List[int]:
    max_r = ws_b.max_row or 0
    out: List[int] = []
    for r in range(ROW_DATA_B, max_r + 1):
        v = ws_b.cell(row=r, column=c_bank).value
        if _coerce_numeric_for_sum(v) is not None:
            out.append(r)
    return out


def _build_header_row_map(ws: Worksheet, header_row: int) -> Dict[int, str]:
    """列号 -> 原始表头文本（仅非空列）。"""
    out: Dict[int, str] = {}
    max_c = ws.max_column or 0
    for c in range(1, max_c + 1):
        raw = ws.cell(row=header_row, column=c).value
        if raw is None or (isinstance(raw, str) and not raw.strip()):
            continue
        out[c] = _cell_text_normalized(raw)
    return out


def _build_b_header_to_col(ws_b: Worksheet) -> Dict[str, int]:
    """表头规范化键 -> 列号（同键取最左列）。"""
    hmap = _build_header_row_map(ws_b, ROW_HEADERS)
    rev: Dict[str, int] = {}
    for c, _text in sorted(hmap.items(), key=lambda x: x[0]):
        key = _header_map_key(hmap[c])
        if key and key not in rev:
            rev[key] = c
    return rev


def _unmerge_cells_overlapping(
    ws: Worksheet, min_row: int, max_row: int, min_col: int, max_col: int
) -> None:
    to_remove: List[str] = []
    for mcr in list(ws.merged_cells.ranges):
        if (
            mcr.min_row <= max_row
            and mcr.max_row >= min_row
            and mcr.min_col <= max_col
            and mcr.max_col >= min_col
        ):
            to_remove.append(str(mcr))
    for s in to_remove:
        ws.unmerge_cells(s)


def _clear_rectangle(ws: Worksheet, r1: int, r2: int, c1: int, c2: int) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _find_footer_row(ws: Worksheet) -> Optional[int]:
    """
    定位表尾行（含「编制」等），行号 1-based。
    约定：B 列含「编制」，且 F 列含「采购总经理」或 I 列含「财务总经理」。
    """
    max_r = min(ws.max_row or 0, 400)
    for r in range(ROW_HEADERS + 1, max_r + 1):
        b = _cell_text_normalized(ws.cell(row=r, column=2).value)
        f = _cell_text_normalized(ws.cell(row=r, column=6).value)
        i = _cell_text_normalized(ws.cell(row=r, column=9).value)
        if "编制" not in b:
            continue
        if "采购总经理" in f or "财务总经理" in i:
            return r
    return None


def _ensure_rows_before_footer(
    ws: Worksheet, footer_row: int, data_start: int, num_data: int
) -> None:
    """在表尾行之前保证有 num_data 个连续空行（从 data_start 起），不足则在 footer_row 处向下插入行。"""
    if num_data <= 0:
        return
    gap = footer_row - data_start
    if gap < num_data:
        ws.insert_rows(footer_row, num_data - gap)


def _apply_data_borders(ws: Worksheet, r1: int, r2: int, c1: int, c2: int) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            cell.border = _DATA_BORDER


def fill_detail_workbook_from_total(detail_path: str, total_path: str) -> List[str]:
    """
    打开明细模板与总表，按 Sheet 遍历填充后保存 detail_path。
    返回每条人类可读日志（跳过原因等）；无异常则至少返回一条汇总。
    """
    logs: List[str] = []
    wb_a = load_workbook(detail_path, read_only=False, data_only=False)
    wb_b = load_workbook(total_path, read_only=False, data_only=True)
    try:
        ws_b = wb_b.worksheets[0]
        b_hdr_to_col = _build_b_header_to_col(ws_b)

        for ws_a in wb_a.worksheets:
            title = ws_a.title
            c_bank = _find_bank_column_for_sheet(ws_b, title)
            if c_bank is None:
                logs.append(f"跳过 Sheet「{title}」：总表第 3 行未匹配到对应银行列。")
                continue

            data_rows = _collect_data_rows_bank_numeric(ws_b, c_bank)
            a_headers = _build_header_row_map(ws_a, ROW_HEADERS)
            if not a_headers:
                logs.append(f"跳过 Sheet「{title}」：第 2 行无表头。")
                continue

            cols_a = sorted(a_headers.keys())
            c_min, c_max = min(cols_a), max(cols_a)
            serial_cols: List[int] = []
            map_ab: Dict[int, Optional[int]] = {}
            for ca in cols_a:
                hk = _header_map_key(a_headers[ca])
                if hk == "序号":
                    serial_cols.append(ca)
                    map_ab[ca] = None
                    continue
                cb = b_hdr_to_col.get(hk)
                map_ab[ca] = cb

            n = len(data_rows)
            footer_row = _find_footer_row(ws_a)
            if footer_row is None:
                logs.append(
                    f"「{title}」：未找到表尾「编制」行，仍从第 {ROW_DATA_A} 行起写入（可能覆盖下方内容）。"
                )
                max_clear_r = max(
                    ws_a.max_row or ROW_DATA_A, ROW_DATA_A + n + 50
                )
                _unmerge_cells_overlapping(
                    ws_a, ROW_DATA_A, max_clear_r, c_min, c_max
                )
                _clear_rectangle(ws_a, ROW_DATA_A, max_clear_r, c_min, c_max)
            else:
                band_end = footer_row - 1
                _unmerge_cells_overlapping(
                    ws_a, ROW_DATA_A, max(band_end, ROW_DATA_A), c_min, c_max
                )
                _clear_rectangle(ws_a, ROW_DATA_A, band_end, c_min, c_max)
                _ensure_rows_before_footer(ws_a, footer_row, ROW_DATA_A, n)

            for idx, r_b in enumerate(data_rows):
                r_a = ROW_DATA_A + idx
                for ca in cols_a:
                    if ca in serial_cols:
                        ws_a.cell(row=r_a, column=ca).value = idx + 1
                        continue
                    cb = map_ab.get(ca)
                    if cb is None:
                        continue
                    ws_a.cell(row=r_a, column=ca).value = ws_b.cell(
                        row=r_b, column=cb
                    ).value

            if n > 0:
                _apply_data_borders(
                    ws_a, ROW_DATA_A, ROW_DATA_A + n - 1, c_min, c_max
                )

            logs.append(
                f"「{title}」：银行列第 {c_bank} 列，已写入 {n} 行。"
            )

        wb_a.save(detail_path)
        logs.insert(0, f"排款明细表已生成：{detail_path}")
    finally:
        wb_b.close()
        wb_a.close()

    return logs
